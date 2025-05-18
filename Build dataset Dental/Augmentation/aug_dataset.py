from datasets import load_dataset
from transformers import MT5Tokenizer, MT5ForConditionalGeneration
from tqdm import tqdm
import pandas as pd
import torch
import os

def augment_and_save_paraphrases(
    dataset_name: str,
    split: str,
    checkpoint: str,
    output_file: str,
    num_return_sequences: int = 10,
    max_length: int = 512,
    device: str = None
):
    """
    Tạo dữ liệu paraphrase từ một dataset và lưu vào file Excel.

    Args:
        dataset_name (str): Tên dataset trên Huggingface datasets.
        split (str): Tên split của dataset (vd: 'train', 'test').
        checkpoint (str): Tên checkpoint model paraphrase huggingface.
        output_file (str): Đường dẫn file Excel để lưu dữ liệu.
        num_return_sequences (int, optional): Số câu paraphrase tạo ra cho mỗi câu gốc. Mặc định 10.
        max_length (int, optional): Độ dài tối đa câu đầu vào và đầu ra. Mặc định 512.
        device (str, optional): Thiết bị chạy mô hình (vd: 'cuda', 'cpu'). Mặc định tự detect.
    """
    # Setup device
    device = device or ("cuda" if torch.cuda.is_available() else "cpu")
    print(f"💻 Đang sử dụng thiết bị: {device}")

    # Load model và tokenizer
    tokenizer = MT5Tokenizer.from_pretrained(checkpoint)
    model = MT5ForConditionalGeneration.from_pretrained(checkpoint).to(device)

    # Load dataset
    dataset = load_dataset(dataset_name, split=split)

    # Kiểm tra và tạo file Excel nếu chưa có
    if not os.path.exists(output_file):
        df_empty = pd.DataFrame(columns=["Câu hỏi", "CoT_Goal", "CoT_Reasoning", "CoT_Justification", "Câu trả lời"])
        df_empty.to_excel(output_file, index=False)

    def paraphrase(text, num_return_sequences=num_return_sequences):
        inputs = tokenizer(text, padding='longest', max_length=max_length, truncation=True, return_tensors='pt')
        inputs = {key: val.to(device) for key, val in inputs.items()}
        outputs = model.generate(
            inputs["input_ids"],
            attention_mask=inputs["attention_mask"],
            max_length=max_length,
            num_return_sequences=num_return_sequences,
            do_sample=True,
            top_p=0.95
        )
        return [tokenizer.decode(o, skip_special_tokens=True) for o in outputs]

    # Mở Excel writer một lần và append dữ liệu theo batch (để tối ưu hơn)
    from openpyxl import load_workbook

    # Đọc số dòng hiện tại để append đúng vị trí
    if os.path.exists(output_file):
        wb = load_workbook(output_file)
        ws = wb.active
        startrow = ws.max_row
        wb.close()
    else:
        startrow = 1

    for example in tqdm(dataset, desc="Đang tạo và lưu dữ liệu paraphrase"):
        question = example["Câu hỏi"]
        try:
            paraphrases = paraphrase(question)
        except Exception as e:
            print(f"Lỗi paraphrase câu: {question}\nLỗi: {e}")
            paraphrases = [question]

        rows = []
        for pq in paraphrases:
            rows.append({
                "Câu hỏi": pq,
                "CoT_Goal": example["CoT_Goal"],
                "CoT_Reasoning": example["CoT_Reasoning"],
                "CoT_Justification": example["CoT_Justification"],
                "Câu trả lời": example["Câu trả lời"]
            })

        df_batch = pd.DataFrame(rows)

        # Append vào file Excel (mode a, startrow tương ứng)
        with pd.ExcelWriter(output_file, mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
            df_batch.to_excel(writer, header=False, index=False, startrow=writer.sheets["Sheet1"].max_row)

    print("✅ Đã lưu toàn bộ dữ liệu paraphrase vào file Excel.")

# Ví dụ gọi hàm:
if __name__ == "__main__":
    augment_and_save_paraphrases(
        dataset_name="NV9523/DentalGPT_SFT",
        split="train",
        checkpoint="chieunq/vietnamese-sentence-paraphase",
        output_file="DentalGPT_SFT_augmented.xlsx",
        num_return_sequences=10,
        max_length=512
    )
